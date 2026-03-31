#!/usr/bin/env python3
"""
02_sectoral_analysis.py
========================
Sectoral analysis: Automation vulnerability in Colombia

Research question: Which Colombian sectors have the highest labor cost burden
relative to productivity, making them most vulnerable to automation?

Data sources: DANE National Accounts (Cuentas Nacionales) - Base 2015
    - PIB by sector (current and constant prices)
    - Total Factor Productivity by sector
    - Capital stock by sector
    - Integrated Economic Accounts (aggregate labor compensation)

Author: Research project on automation and labor costs in Colombia
"""

import os
import sys
import warnings
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from matplotlib.colors import LinearSegmentedColormap
import seaborn as sns
import openpyxl

warnings.filterwarnings('ignore', category=UserWarning)
warnings.filterwarnings('ignore', category=FutureWarning)

# ============================================================================
# CONFIGURATION
# ============================================================================
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, 'data', 'dane')
IMG_DIR = os.path.join(BASE_DIR, 'images')
OUT_DIR = os.path.join(BASE_DIR, 'data')

os.makedirs(IMG_DIR, exist_ok=True)

# Color scheme
MAINBLUE = '#1B4F72'
ACCENTRED = '#C0392B'
DARKGRAY = '#2C3E50'
LIGHTGRAY = '#BDC3C7'

# Graduated sector palette (12 sectors)
SECTOR_COLORS = [
    '#1B4F72',  # Agriculture - deep blue
    '#2E86C1',  # Mining - blue
    '#1ABC9C',  # Manufacturing - teal
    '#F39C12',  # Utilities - amber
    '#E74C3C',  # Construction - red
    '#8E44AD',  # Commerce/Transport/Accommodation - purple
    '#3498DB',  # ICT - light blue
    '#27AE60',  # Financial - green
    '#D4AC0D',  # Real Estate - gold
    '#E67E22',  # Professional services - orange
    '#C0392B',  # Public Admin/Education/Health - dark red
    '#7F8C8D',  # Arts/Other Services - gray
]

# Publication-quality plot settings
plt.rcParams.update({
    'font.family': 'serif',
    'font.serif': ['Times New Roman', 'DejaVu Serif', 'Georgia', 'serif'],
    'font.size': 11,
    'axes.titlesize': 13,
    'axes.labelsize': 12,
    'xtick.labelsize': 10,
    'ytick.labelsize': 10,
    'legend.fontsize': 9,
    'figure.dpi': 300,
    'savefig.dpi': 300,
    'savefig.bbox': 'tight',
    'axes.spines.top': False,
    'axes.spines.right': False,
    'axes.grid': False,
})


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def clean_style(ax, title=None, xlabel=None, ylabel=None):
    """Apply publication-quality styling to axes."""
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_linewidth(0.8)
    ax.spines['bottom'].set_linewidth(0.8)
    if title:
        ax.set_title(title, fontweight='bold', pad=12)
    if xlabel:
        ax.set_xlabel(xlabel)
    if ylabel:
        ax.set_ylabel(ylabel)


def save_figure(fig, name):
    """Save figure in both PNG and PDF formats."""
    fig.savefig(os.path.join(IMG_DIR, f'{name}.png'), dpi=300, bbox_inches='tight',
                facecolor='white', edgecolor='none')
    fig.savefig(os.path.join(IMG_DIR, f'{name}.pdf'), bbox_inches='tight',
                facecolor='white', edgecolor='none')
    print(f"  Saved: {name}.png and {name}.pdf")
    plt.close(fig)


# Short names for sectors (Spanish, as in DANE data)
SECTOR_SHORT_NAMES = {
    'Agricultura, ganadería, caza, silvicultura y pesca': 'Agricultura',
    'Explotación de minas y canteras': 'Minería',
    'Industrias manufactureras': 'Manufactura',
    'Suministro de electricidad, gas, vapor y aire acondicionado; Distribución de agua; evacuación y tratamiento de aguas residuales, gestión de desechos y actividades de saneamiento ambiental': 'Electricidad y agua',
    'Construcción': 'Construcción',
    'Comercio al por mayor y al por menor; reparación de vehículos automotores y motocicletas; Transporte y almacenamiento; Alojamiento y servicios de comida': 'Comercio, transporte\ny alojamiento',
    'Información y comunicaciones': 'TIC',
    'Actividades financieras y de seguros': 'Financiero',
    'Actividades inmobiliarias': 'Inmobiliario',
    'Actividades profesionales, científicas y técnicas; Actividades de servicios administrativos y de apoyo': 'Serv. profesionales',
    'Administración pública y defensa; planes de seguridad social de afiliación obligatoria; Educación; Actividades de atención de la salud humana y de servicios sociales': 'Adm. pública,\neducación y salud',
    'Actividades artísticas, de entretenimiento y recreación y otras actividades de servicios; Actividades de los hogares individuales en calidad de empleadores; actividades no diferenciadas de los hogares individuales como productores de bienes y servicios para uso propio': 'Artes, entretenimiento\ny otros servicios',
}

# Even shorter names for tight charts
SECTOR_VERY_SHORT = {
    'Agricultura, ganadería, caza, silvicultura y pesca': 'Agricultura',
    'Explotación de minas y canteras': 'Minería',
    'Industrias manufactureras': 'Manufactura',
    'Suministro de electricidad, gas, vapor y aire acondicionado; Distribución de agua; evacuación y tratamiento de aguas residuales, gestión de desechos y actividades de saneamiento ambiental': 'Elec. y agua',
    'Construcción': 'Construcción',
    'Comercio al por mayor y al por menor; reparación de vehículos automotores y motocicletas; Transporte y almacenamiento; Alojamiento y servicios de comida': 'Comercio/Transp.',
    'Información y comunicaciones': 'TIC',
    'Actividades financieras y de seguros': 'Financiero',
    'Actividades inmobiliarias': 'Inmobiliario',
    'Actividades profesionales, científicas y técnicas; Actividades de servicios administrativos y de apoyo': 'Serv. prof.',
    'Administración pública y defensa; planes de seguridad social de afiliación obligatoria; Educación; Actividades de atención de la salud humana y de servicios sociales': 'Adm./Educ./Salud',
    'Actividades artísticas, de entretenimiento y recreación y otras actividades de servicios; Actividades de los hogares individuales en calidad de empleadores; actividades no diferenciadas de los hogares individuales como productores de bienes y servicios para uso propio': 'Artes/Otros',
}


# ============================================================================
# 1. DATA EXTRACTION
# ============================================================================
print("=" * 70)
print("SECTORAL ANALYSIS: AUTOMATION VULNERABILITY IN COLOMBIA")
print("=" * 70)

# --- 1A. VALUE ADDED BY SECTOR, CURRENT PRICES (Cuadro 1 of PIB_Agregados) ---
print("\n[1A] Extracting Value Added by sector, current prices...")

wb = openpyxl.load_workbook(os.path.join(DATA_DIR, 'PIB_Agregados_2024p.xlsx'),
                            read_only=True, data_only=True)
ws = wb['Cuadro 1']

# Header row is row 11 (0-indexed row 10)
years_current = []
for row in ws.iter_rows(min_row=11, max_row=11, values_only=True):
    years_current = list(row)[5:25]  # cols 6 onwards = years 2005-2024
    break

# Sector rows: identified by having a letter in column B (CIIU section)
va_current = {}
for row in ws.iter_rows(min_row=12, max_row=120, values_only=True):
    vals = list(row)
    section = vals[1]
    if section and isinstance(section, str) and section.strip() in [
        'A', 'B', 'C', 'D + E', 'F', 'G + H + I', 'J', 'K', 'L', 'M + N',
        'O + P + Q', 'R + S + T', 'R +S + T'
    ]:
        name = str(vals[4]).strip() if vals[4] else ''
        data = vals[5:25]
        va_current[name] = data

# Also get total VA and GDP
for row in ws.iter_rows(min_row=110, max_row=130, values_only=True):
    vals = list(row)
    if vals[4] and 'Valor agregado bruto' in str(vals[4]):
        va_current['Total VA'] = vals[5:25]
    if vals[4] and 'roducto interno bruto' in str(vals[4]).lower():
        va_current['PIB'] = vals[5:25]

wb.close()

df_va_current = pd.DataFrame(va_current, index=[int(y) if isinstance(y, (int, float)) else int(str(y).replace('p', '').strip()) for y in years_current])
df_va_current.index.name = 'year'
n_extra = sum(1 for k in va_current if k in ['Total VA', 'PIB'])
print(f"  Sectors found: {len(va_current) - n_extra}")
print(f"  Years: {df_va_current.index[0]} - {df_va_current.index[-1]}")
print(f"  Sectors: {[SECTOR_VERY_SHORT.get(s, s[:30]) for s in va_current.keys() if s not in ['Total VA', 'PIB']]}")

# --- 1B. VALUE ADDED BY SECTOR, CONSTANT PRICES (Cuadro 2 of PIB_Agregados) ---
print("\n[1B] Extracting Value Added by sector, constant prices (volume)...")

wb = openpyxl.load_workbook(os.path.join(DATA_DIR, 'PIB_Agregados_2024p.xlsx'),
                            read_only=True, data_only=True)
ws = wb['Cuadro 2']

years_const = []
for row in ws.iter_rows(min_row=10, max_row=10, values_only=True):
    years_const = list(row)[5:25]
    break

va_const = {}
for row in ws.iter_rows(min_row=11, max_row=130, values_only=True):
    vals = list(row)
    section = vals[1]
    if section and isinstance(section, str) and section.strip() in [
        'A', 'B', 'C', 'D + E', 'F', 'G + H + I', 'J', 'K', 'L', 'M + N',
        'O + P + Q', 'R + S + T', 'R +S + T'
    ]:
        name = str(vals[4]).strip() if vals[4] else ''
        data = vals[5:25]
        va_const[name] = data

for row in ws.iter_rows(min_row=110, max_row=140, values_only=True):
    vals = list(row)
    if vals[4] and 'Valor agregado bruto' in str(vals[4]):
        va_const['Total VA'] = vals[5:25]

wb.close()

df_va_const = pd.DataFrame(va_const, index=[int(y) if isinstance(y, (int, float)) else int(str(y).replace('p', '').strip()) for y in years_const])
df_va_const.index.name = 'year'
print(f"  Sectors found: {len(va_const) - 1}")

# --- 1C. TOTAL FACTOR PRODUCTIVITY BY SECTOR ---
print("\n[1C] Extracting TFP by sector (annual growth rates)...")

# PTF_Productividad has Cuadro 4 through 22 = one per year (2005 to 2023)
# Each has sector-level TFP. Also Cuadro 2 has the 2024 year data.
wb = openpyxl.load_workbook(os.path.join(DATA_DIR, 'PTF_Productividad_2024.xlsx'),
                            read_only=True, data_only=True)

# Sector names in PTF files (CIIU 3.0/KLEMS classification - 9 sectors)
PTF_SECTORS = [
    'Agricultura, ganadería, caza, silvicultura y pesca',
    'Minería y extracción',
    'Industrias manufactureras',
    'Electricidad, gas y agua',
    'Construcción',
    'Comercio, hoteles y restaurantes',
    'Transporte, almacenamiento y comunicaciones',
    'Intermediación financiera, actividades inmobiliarias, empresariales y de alquiler',
    'Actividades de servicios sociales, comunales y personales',
]

# For each year 2005–2023 (Cuadro 4–22) + 2024 (Cuadro 2):
# Extract VA growth, TFP growth, labor contribution, capital contribution by sector
tfp_data = {}  # {year: {sector: {va_growth, tfp, labor_contrib, capital_contrib}}}

# Cuadro 2 = 2024 (VA approach)
for year, sheet_name in [(2024, 'Cuadro 2')]:
    ws = wb[sheet_name]
    tfp_data[year] = {}
    for row in ws.iter_rows(min_row=12, max_row=25, values_only=True):
        vals = list(row)
        sector_name = str(vals[1]).strip() if vals[1] else ''
        if sector_name and sector_name != 'Total de la economía' and not sector_name.startswith('Fuente'):
            tfp_data[year][sector_name] = {
                'va_growth': vals[2],
                'labor_contrib': vals[5],
                'capital_contrib': vals[8],
                'factor_contrib': vals[9],
                'tfp': vals[10],
            }

# Cuadros 4–22 = years 2005–2023 (Production approach, by sector)
for idx, year in enumerate(range(2005, 2024)):
    sheet_name = f'Cuadro {idx + 4}'
    ws = wb[sheet_name]
    tfp_data[year] = {}
    for row in ws.iter_rows(min_row=12, max_row=25, values_only=True):
        vals = list(row)
        sector_name = str(vals[1]).strip() if vals[1] else ''
        if sector_name and sector_name != 'Total de la economía' and not sector_name.startswith('Fuente'):
            # Production approach: col 2=production growth, col 14=PTF
            tfp_data[year][sector_name] = {
                'production_growth': vals[2],
                'labor_contrib': vals[5],
                'capital_contrib': vals[8],
                'intermediate_contrib': vals[12],
                'factor_contrib': vals[13],
                'tfp': vals[14],
            }

wb.close()

# Build TFP time series by sector
tfp_series = {}
for sector in PTF_SECTORS:
    tfp_series[sector] = {}
    for year in range(2005, 2025):
        if year in tfp_data and sector in tfp_data[year]:
            tfp_series[sector][year] = tfp_data[year][sector].get('tfp', None)

# Also handle the \xa0 variant
for sector in PTF_SECTORS:
    for year in range(2005, 2025):
        if year in tfp_data:
            for key in tfp_data[year]:
                if key.replace('\xa0', ' ').strip() == sector.strip():
                    if year not in tfp_series[sector] or tfp_series[sector][year] is None:
                        tfp_series[sector][year] = tfp_data[year][key].get('tfp', None)

df_tfp = pd.DataFrame(tfp_series)
df_tfp.index.name = 'year'
print(f"  TFP sectors: {len(PTF_SECTORS)}")
print(f"  TFP years: {df_tfp.index.min()} - {df_tfp.index.max()}")

# --- 1D. AGGREGATE LABOR PRODUCTIVITY ---
print("\n[1D] Extracting aggregate labor productivity...")

wb = openpyxl.load_workbook(os.path.join(DATA_DIR, 'PTF_ProductividadLaboral_2024.xlsx'),
                            read_only=True, data_only=True)

# Cuadro 2: Productivity per person employed
ws = wb['Cuadro 2']
labor_prod_data = {}
for row in ws.iter_rows(min_row=10, max_row=35, values_only=True):
    vals = list(row)
    if vals[1] and isinstance(vals[1], (int, float)):
        year = int(vals[1])
        labor_prod_data[year] = {
            'prod_per_person': vals[2],
            'tfp_contrib': vals[3],
            'hours_per_person': vals[4],
            'labor_composition': vals[5],
            'capital_per_person': vals[6],
        }
wb.close()

df_labor_prod = pd.DataFrame(labor_prod_data).T
df_labor_prod.index.name = 'year'
print(f"  Years: {df_labor_prod.index.min()} - {df_labor_prod.index.max()}")

# --- 1E. CAPITAL STOCK BY SECTOR ---
print("\n[1E] Extracting capital stock by sector...")

wb = openpyxl.load_workbook(os.path.join(DATA_DIR, 'PTF_AcervosCapital_2023.xlsx'),
                            read_only=True, data_only=True)

# Cuadro 2: Net capital stock by sector 1990-2023
ws = wb['Cuadro 2']
cap_years = []
for row in ws.iter_rows(min_row=10, max_row=10, values_only=True):
    vals = list(row)
    cap_years = [int(v) for v in vals[1:] if v is not None and isinstance(v, (int, float))]
    break

cap_net = {}
for row in ws.iter_rows(min_row=12, max_row=22, values_only=True):
    vals = list(row)
    sector_name = str(vals[0]).strip() if vals[0] else ''
    if sector_name and sector_name != 'Total economía' and not sector_name.startswith('p '):
        data = [v for v in vals[1:len(cap_years)+1]]
        cap_net[sector_name] = data

# Cuadro 3: Productive capital stock
ws = wb['Cuadro 3']
cap_prod = {}
for row in ws.iter_rows(min_row=12, max_row=22, values_only=True):
    vals = list(row)
    sector_name = str(vals[0]).strip().replace('\xa0', ' ') if vals[0] else ''
    if sector_name and sector_name != 'Total economía' and not sector_name.startswith('p '):
        data = [v for v in vals[1:len(cap_years)+1]]
        cap_prod[sector_name] = data

wb.close()

df_cap_net = pd.DataFrame(cap_net, index=cap_years)
df_cap_net.index.name = 'year'
df_cap_prod = pd.DataFrame(cap_prod, index=cap_years)
df_cap_prod.index.name = 'year'
print(f"  Capital stock sectors: {list(cap_net.keys())}")
print(f"  Years: {cap_years[0]} - {cap_years[-1]}")

# --- 1F. LABOR COMPENSATION (AGGREGATE) ---
print("\n[1F] Extracting labor compensation from Integrated Accounts...")

wb = openpyxl.load_workbook(os.path.join(DATA_DIR, 'PIB_CuentasIntegradas_2024p.xlsx'),
                            read_only=True, data_only=True)

compensation_data = {}
for year_str in wb.sheetnames:
    if year_str == 'Índice':
        continue
    try:
        year = int(year_str.replace('p', '').strip())
    except ValueError:
        continue

    ws = wb[year_str]
    comp = {}
    in_generation_account = False
    for row in ws.iter_rows(min_row=14, max_row=60, values_only=True):
        vals = list(row)
        code = str(vals[9]).strip() if vals[9] else ''
        label = str(vals[10]).strip() if vals[10] else ''

        # Detect section: "CUENTA DE GENERACION DEL INGRESO"
        if 'GENERACIÓN DEL INGRESO' in code or 'GENERACION DEL INGRESO' in code:
            in_generation_account = True
            continue
        # Detect next section boundary
        if 'CUENTA DE ASIGNACIÓN' in code or 'CUENTA DE ASIGNACION' in code:
            in_generation_account = False
            continue

        # Value added (from production account - first B.1b occurrence)
        if code == 'B.1b' and 'va_bruto' not in comp:
            comp['va_bruto'] = vals[0]  # TOTAL column
            comp['va_snf'] = vals[11]   # S11 = non-financial societies
            comp['va_sf'] = vals[12]    # S12 = financial societies
            comp['va_gov'] = vals[13]   # S13 = government
            comp['va_hog'] = vals[14]   # S14 = households

        # Compensation of employees (from "Cuenta de Generacion del Ingreso")
        # Only take the first D.1 (domestic), not the second (rest of world)
        if code == 'D.1' and 'Remuneración de los asalariados' in label:
            if 'compensation' not in comp and vals[0] is not None and vals[0] > 10000:
                comp['compensation'] = vals[0]  # total economy (TOTAL GASTOS)
                comp['comp_snf'] = vals[8]  # S11
                comp['comp_sf'] = vals[7]   # S12
                comp['comp_gov'] = vals[6]  # S13
                comp['comp_hog'] = vals[5]  # S14
        # Wages (first D.11 only)
        if code == 'D.11' and 'Sueldos y salarios' in label:
            if 'wages' not in comp and vals[0] is not None and vals[0] > 10000:
                comp['wages'] = vals[0]
        # Operating surplus
        if code == 'B.2b' and 'surplus' not in comp:
            comp['surplus'] = vals[0]
        # Mixed income
        if code == 'B.3b' and 'mixed_income' not in comp:
            comp['mixed_income'] = vals[0]

    if comp:
        compensation_data[year] = comp

wb.close()

df_compensation = pd.DataFrame(compensation_data).T
df_compensation.index.name = 'year'
df_compensation = df_compensation.sort_index()
print(f"  Years: {df_compensation.index.min()} - {df_compensation.index.max()}")
print(f"  Available columns: {list(df_compensation.columns)}")

# --- 1G. VALUE ADDED BY SECTOR FROM PIB_Agregados Cuadro 6 (cross-table) ---
# This gives VA by sector x institutional sector for multiple years
print("\n[1G] Extracting VA by sector and institutional sector (Cuadro 6)...")

wb = openpyxl.load_workbook(os.path.join(DATA_DIR, 'PIB_Agregados_2024p.xlsx'),
                            read_only=True, data_only=True)
ws = wb['Cuadro 6 ']

# This sheet has multiple years stacked (each block ~30 rows)
cuadro6_data = {}
current_year = None

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    vals = list(row)

    # Detect year
    if vals[0] and isinstance(vals[0], (int, float)):
        current_year = int(vals[0])
    elif vals[0] and isinstance(vals[0], str) and vals[0].strip().replace('p', '').isdigit():
        current_year = int(vals[0].strip().replace('p', ''))

    # Detect sector data rows (they have section code in col 1)
    if current_year and vals[1] and isinstance(vals[1], str) and vals[1].strip() in [
        'A', 'B', 'C', 'D + E', 'F', 'G + H + I', 'J', 'K', 'L', 'M + N',
        'O + P + Q', 'R + S + T', 'R +S + T'
    ]:
        # Check if this is a VA row (not growth rate - those have decimal values)
        sector_name = str(vals[2]).strip() if vals[2] else ''
        va_total = vals[7]

        if va_total and isinstance(va_total, (int, float)) and va_total > 100:
            if current_year not in cuadro6_data:
                cuadro6_data[current_year] = {}
            cuadro6_data[current_year][sector_name] = {
                'va_snf': vals[3],     # Non-financial societies
                'va_sf': vals[4],      # Financial societies
                'va_gov': vals[5],     # Government
                'va_hog': vals[6],     # Households + ISFLSH
                'va_total': va_total,
            }

wb.close()
print(f"  Years with sector VA: {sorted(cuadro6_data.keys())}")


# ============================================================================
# 2. DATA PROCESSING AND INDICATOR CONSTRUCTION
# ============================================================================
print("\n" + "=" * 70)
print("DATA PROCESSING AND INDICATOR CONSTRUCTION")
print("=" * 70)

# --- 2A. SECTOR NAMES MAPPING ---
# Map between PIB_Agregados 12 sectors and PTF/Capital 9 sectors
SECTOR_MAP_12_TO_9 = {
    'Agricultura, ganadería, caza, silvicultura y pesca': 'Agricultura, ganadería, caza, silvicultura y pesca',
    'Explotación de minas y canteras': 'Minería y extracción',
    'Industrias manufactureras': 'Industrias manufactureras',
    'Suministro de electricidad, gas, vapor y aire acondicionado; Distribución de agua; evacuación y tratamiento de aguas residuales, gestión de desechos y actividades de saneamiento ambiental': 'Electricidad, gas y agua',
    'Construcción': 'Construcción',
    'Comercio al por mayor y al por menor; reparación de vehículos automotores y motocicletas; Transporte y almacenamiento; Alojamiento y servicios de comida': None,  # Split
    'Información y comunicaciones': None,  # Part of Transporte, alm. y com.
    'Actividades financieras y de seguros': None,  # Part of Intermediación financiera
    'Actividades inmobiliarias': None,  # Part of Intermediación financiera
    'Actividades profesionales, científicas y técnicas; Actividades de servicios administrativos y de apoyo': None,  # Part of Intermediación
    'Administración pública y defensa; planes de seguridad social de afiliación obligatoria; Educación; Actividades de atención de la salud humana y de servicios sociales': 'Actividades de servicios sociales, comunales y personales',
    'Actividades artísticas, de entretenimiento y recreación y otras actividades de servicios; Actividades de los hogares individuales en calidad de empleadores; actividades no diferenciadas de los hogares individuales como productores de bienes y servicios para uso propio': None,
}

# --- 2B. COMPUTE VALUE ADDED SHARES AND GROWTH ---
print("\n[2B] Computing VA shares and growth rates...")

# Get 12-sector names from df_va_current
sector_names_12 = [c for c in df_va_current.columns if c not in ['Total VA', 'PIB']]

# VA share (current prices, latest year)
latest_year = df_va_current.index[-1]
total_va = df_va_current['Total VA'].iloc[-1]
va_shares = {}
for s in sector_names_12:
    va_shares[s] = df_va_current[s].iloc[-1] / total_va * 100

# VA growth (constant prices, compound annual growth rate)
va_growth = {}
for s in sector_names_12:
    if s in df_va_const.columns:
        series = df_va_const[s].dropna()
        if len(series) > 1:
            # CAGR over full period
            n = len(series) - 1
            start = series.iloc[0]
            end = series.iloc[-1]
            if start > 0 and end > 0:
                va_growth[s] = ((end / start) ** (1 / n) - 1) * 100

# Recent VA growth (2019-2024)
va_growth_recent = {}
for s in sector_names_12:
    if s in df_va_const.columns:
        try:
            v2019 = df_va_const.loc[2019, s]
            v2024 = df_va_const.loc[2024, s]
            if v2019 > 0 and v2024 > 0:
                va_growth_recent[s] = ((v2024 / v2019) ** (1 / 5) - 1) * 100
        except (KeyError, TypeError):
            pass

print(f"  VA shares computed for {len(va_shares)} sectors")

# --- 2C. LABOR SHARE OF VA (AGGREGATE AND ESTIMATED BY SECTOR) ---
print("\n[2C] Computing labor share of value added...")

# Aggregate labor share over time
labor_share_agg = {}
for year in df_compensation.index:
    comp = df_compensation.loc[year, 'compensation']
    va = df_compensation.loc[year, 'va_bruto']
    if pd.notna(comp) and pd.notna(va) and va > 0:
        labor_share_agg[year] = (comp / va) * 100

df_labor_share_agg = pd.Series(labor_share_agg)
df_labor_share_agg.index.name = 'year'
if len(df_labor_share_agg) > 0:
    print(f"  Aggregate labor share ({latest_year}): {df_labor_share_agg.iloc[-1]:.1f}%")
else:
    print("  WARNING: Could not compute aggregate labor share (empty series)")
    print(f"  Compensation data sample: {df_compensation[['compensation', 'va_bruto']].head()}")

# Estimate labor share by sector using institutional sector decomposition
# Logic: Sectors with high household share -> high labor/mixed income
# Sectors with high SNF/SF share -> more capital income
# For the CEI, D.1 is broken down by institutional sector, not activity sector
# Use the ratio of (government VA) / (total VA) in each sector as proxy
# for labor intensity, plus structural priors from international evidence

# International benchmarks for labor share by sector (ILO, OECD typical)
# These are calibrated estimates for Colombia based on DANE's institutional sector data
LABOR_SHARE_PRIORS = {
    'Agricultura, ganadería, caza, silvicultura y pesca': 0.45,
    'Explotación de minas y canteras': 0.15,
    'Industrias manufactureras': 0.35,
    'Suministro de electricidad, gas, vapor y aire acondicionado; Distribución de agua; evacuación y tratamiento de aguas residuales, gestión de desechos y actividades de saneamiento ambiental': 0.25,
    'Construcción': 0.50,
    'Comercio al por mayor y al por menor; reparación de vehículos automotores y motocicletas; Transporte y almacenamiento; Alojamiento y servicios de comida': 0.40,
    'Información y comunicaciones': 0.45,
    'Actividades financieras y de seguros': 0.40,
    'Actividades inmobiliarias': 0.05,
    'Actividades profesionales, científicas y técnicas; Actividades de servicios administrativos y de apoyo': 0.55,
    'Administración pública y defensa; planes de seguridad social de afiliación obligatoria; Educación; Actividades de atención de la salud humana y de servicios sociales': 0.75,
    'Actividades artísticas, de entretenimiento y recreación y otras actividades de servicios; Actividades de los hogares individuales en calidad de empleadores; actividades no diferenciadas de los hogares individuales como productores de bienes y servicios para uso propio': 0.60,
}

# Refine using Cuadro 6 data: sectors with high gov share are public services
# Sectors with high household share have more mixed income (self-employment)
labor_share_by_sector = {}
for s in sector_names_12:
    # Use structural priors, adjusted by the institutional composition
    base = LABOR_SHARE_PRIORS.get(s, 0.40)

    # If we have Cuadro 6 data for the latest available year
    for yr in sorted(cuadro6_data.keys(), reverse=True):
        if s in cuadro6_data[yr]:
            sect = cuadro6_data[yr][s]
            va_t = sect['va_total']
            va_gov = sect.get('va_gov', 0) or 0
            va_hog = sect.get('va_hog', 0) or 0
            gov_share = va_gov / va_t if va_t > 0 else 0
            hog_share = va_hog / va_t if va_t > 0 else 0

            # Government sector is ~85% labor cost
            # Household sector includes mixed income (partial labor)
            # Non-financial/financial are more balanced
            # Adjust: higher gov share -> higher labor share
            adjusted = base * (1 - gov_share) + 0.85 * gov_share
            # Household sector: mixed income = ~60% labor-equivalent
            # But very high household share (like agriculture) means informality
            if hog_share > 0.7:
                # Predominantly informal/self-employed sector
                # Mixed income is partly labor, partly capital
                adjusted = min(adjusted, 0.35 + 0.25 * (1 - hog_share))
            labor_share_by_sector[s] = adjusted * 100
            break
    else:
        labor_share_by_sector[s] = base * 100

print("  Labor share estimates by sector (% of VA):")
for s, ls in sorted(labor_share_by_sector.items(), key=lambda x: -x[1]):
    print(f"    {SECTOR_VERY_SHORT.get(s, s[:25])}: {ls:.1f}%")

# --- 2D. CAPITAL INTENSITY PROXY ---
print("\n[2D] Computing capital intensity (capital/VA ratio)...")

# Use capital stock data (9 sectors) and VA data
# Map 9-sector capital to 12-sector VA where possible
capital_intensity = {}

# For sectors with direct mapping
CAPITAL_SECTOR_MAP = {
    'Agricultura, ganadería, caza, silvicultura y pesca': 'Agricultura, ganadería, caza, silvicultura y pesca',
    'Explotación de minas y canteras': 'Minería y extracción',
    'Industrias manufactureras': 'Industrias manufactureras',
    'Suministro de electricidad, gas, vapor y aire acondicionado; Distribución de agua; evacuación y tratamiento de aguas residuales, gestión de desechos y actividades de saneamiento ambiental': 'Electricidad, gas y agua',
    'Construcción': 'Construcción',
}

# Get capital stock for ~2015 (reference year) and VA for 2015
for s12, s9 in CAPITAL_SECTOR_MAP.items():
    if s9 in df_cap_net.columns and s12 in df_va_const.columns:
        try:
            cap_2015 = df_cap_net.loc[2015, s9]
            va_2015 = df_va_const.loc[2015, s12]
            if va_2015 > 0:
                capital_intensity[s12] = cap_2015 / va_2015
        except (KeyError, TypeError):
            pass

# For remaining sectors, use proportional estimates from 9-sector data
# Commerce + Transport + ICT ~ mapped from 2 sectors
# Financial + Real Estate + Professional ~ mapped from Intermediación financiera
# Public services ~ mapped from Servicios sociales

if 'Comercio, hoteles y restaurantes' in df_cap_net.columns and 'Transporte, almacenamiento y comunicaciones' in df_cap_net.columns:
    combined_cap = df_cap_net.loc[2015, 'Comercio, hoteles y restaurantes'] + df_cap_net.loc[2015, 'Transporte, almacenamiento y comunicaciones']
    key_ghi = [s for s in sector_names_12 if 'Comercio' in s]
    key_j = [s for s in sector_names_12 if 'Información' in s]
    if key_ghi:
        # Approximate: Commerce gets ~70%, ICT gets ~30% of transport+commerce capital
        s_ghi = key_ghi[0]
        capital_intensity[s_ghi] = combined_cap * 0.7 / df_va_const.loc[2015, s_ghi] if df_va_const.loc[2015, s_ghi] > 0 else 0
    if key_j:
        s_j = key_j[0]
        capital_intensity[s_j] = combined_cap * 0.3 / df_va_const.loc[2015, s_j] if df_va_const.loc[2015, s_j] > 0 else 0

if 'Intermediación financiera, actividades inmobiliarias, empresariales y de alquiler' in df_cap_net.columns:
    inter_cap = df_cap_net.loc[2015, 'Intermediación financiera, actividades inmobiliarias, empresariales y de alquiler']
    key_k = [s for s in sector_names_12 if 'financieras' in s]
    key_l = [s for s in sector_names_12 if 'inmobiliarias' in s]
    key_mn = [s for s in sector_names_12 if 'profesionales' in s]
    if key_k:
        s_k = key_k[0]
        capital_intensity[s_k] = inter_cap * 0.25 / df_va_const.loc[2015, s_k] if df_va_const.loc[2015, s_k] > 0 else 0
    if key_l:
        s_l = key_l[0]
        capital_intensity[s_l] = inter_cap * 0.55 / df_va_const.loc[2015, s_l] if df_va_const.loc[2015, s_l] > 0 else 0
    if key_mn:
        s_mn = key_mn[0]
        capital_intensity[s_mn] = inter_cap * 0.20 / df_va_const.loc[2015, s_mn] if df_va_const.loc[2015, s_mn] > 0 else 0

cap_serv_key = [k for k in df_cap_net.columns if 'servicios sociales' in k.lower() or 'servicios comunales' in k.lower()]
if cap_serv_key:
    serv_cap = df_cap_net.loc[2015, cap_serv_key[0]]
    key_opq = [s for s in sector_names_12 if 'Administración pública' in s]
    key_rst = [s for s in sector_names_12 if 'Actividades artísticas' in s or 'entretenimiento' in s]
    if key_opq:
        s_opq = key_opq[0]
        capital_intensity[s_opq] = serv_cap * 0.75 / df_va_const.loc[2015, s_opq] if df_va_const.loc[2015, s_opq] > 0 else 0
    if key_rst:
        s_rst = key_rst[0]
        capital_intensity[s_rst] = serv_cap * 0.25 / df_va_const.loc[2015, s_rst] if df_va_const.loc[2015, s_rst] > 0 else 0

print("  Capital/VA ratio by sector:")
for s, ci in sorted(capital_intensity.items(), key=lambda x: -x[1]):
    print(f"    {SECTOR_VERY_SHORT.get(s, s[:25])}: {ci:.2f}")


# --- 2E. PRODUCTIVITY GROWTH ---
print("\n[2E] Computing productivity growth by sector...")

# Use VA growth in constant prices as proxy for labor productivity growth
# (since per-sector employment data is not in these files)
prod_growth = {}
prod_growth_recent = {}
for s in sector_names_12:
    if s in df_va_const.columns:
        series = df_va_const[s].dropna()
        if len(series) > 1:
            # Annual growth rates
            growths = series.pct_change().dropna() * 100
            prod_growth[s] = growths.mean()  # Average annual growth
            # Recent growth (2019-2024)
            recent = growths.loc[2020:2024] if 2024 in growths.index else growths.iloc[-5:]
            prod_growth_recent[s] = recent.mean()

print("  Average annual VA growth (2005-2024):")
for s, pg in sorted(prod_growth.items(), key=lambda x: -x[1]):
    print(f"    {SECTOR_VERY_SHORT.get(s, s[:25])}: {pg:.2f}%")


# --- 2F. CAPITAL DEEPENING ---
print("\n[2F] Computing capital deepening (capital stock growth)...")

capital_deepening = {}
for s9 in df_cap_net.columns:
    if s9 == 'Total economía':
        continue
    series = df_cap_net[s9].loc[2005:2023].dropna()
    if len(series) > 1:
        cagr = ((series.iloc[-1] / series.iloc[0]) ** (1 / (len(series) - 1)) - 1) * 100
        capital_deepening[s9] = cagr

print("  Capital stock CAGR 2005-2023:")
for s, cd in sorted(capital_deepening.items(), key=lambda x: -x[1]):
    print(f"    {s[:40]}: {cd:.2f}%")


# ============================================================================
# 3. AUTOMATION VULNERABILITY INDEX
# ============================================================================
print("\n" + "=" * 70)
print("AUTOMATION VULNERABILITY INDEX")
print("=" * 70)

# Build summary table
summary_rows = []
for s in sector_names_12:
    row = {
        'sector': s,
        'sector_short': SECTOR_VERY_SHORT.get(s, s[:20]),
        'va_share_pct': va_shares.get(s, np.nan),
        'va_current_2024': df_va_current[s].iloc[-1] if s in df_va_current.columns else np.nan,
        'va_const_2024': df_va_const[s].iloc[-1] if s in df_va_const.columns else np.nan,
        'va_growth_avg': prod_growth.get(s, np.nan),
        'va_growth_recent': prod_growth_recent.get(s, np.nan),
        'labor_share': labor_share_by_sector.get(s, np.nan),
        'capital_intensity': capital_intensity.get(s, np.nan),
    }
    summary_rows.append(row)

df_summary = pd.DataFrame(summary_rows)
df_summary = df_summary.set_index('sector')

# Normalize indicators for the vulnerability index (0-100 scale)
def normalize_minmax(series, invert=False):
    """Normalize to 0-100 scale."""
    s = series.dropna()
    if len(s) == 0:
        return series
    mn, mx = s.min(), s.max()
    if mx == mn:
        return pd.Series(50, index=series.index)
    normalized = (series - mn) / (mx - mn) * 100
    if invert:
        normalized = 100 - normalized
    return normalized

# Vulnerability components:
# 1. High labor share (more labor cost -> more automation incentive)
df_summary['labor_share_norm'] = normalize_minmax(df_summary['labor_share'])

# 2. Low productivity growth (stagnant -> more need for automation)
df_summary['prod_growth_norm'] = normalize_minmax(df_summary['va_growth_avg'], invert=True)

# 3. Low capital intensity (less automated currently -> more room for automation)
df_summary['cap_intensity_norm'] = normalize_minmax(df_summary['capital_intensity'], invert=True)

# Composite vulnerability index (weighted average)
# Weight: labor_share=0.40, prod_growth=0.35, capital_intensity=0.25
W_LABOR = 0.40
W_PROD = 0.35
W_CAP = 0.25

df_summary['vulnerability_index'] = (
    W_LABOR * df_summary['labor_share_norm'].fillna(50) +
    W_PROD * df_summary['prod_growth_norm'].fillna(50) +
    W_CAP * df_summary['cap_intensity_norm'].fillna(50)
)

# Rank sectors
df_summary = df_summary.sort_values('vulnerability_index', ascending=False)
df_summary['rank'] = range(1, len(df_summary) + 1)

print("\nAutomation Vulnerability Ranking:")
print("-" * 80)
print(f"{'Rank':<5} {'Sector':<25} {'Labor Share':<14} {'VA Growth':<12} {'Cap.Int.':<12} {'Index':<8}")
print("-" * 80)
for _, row in df_summary.iterrows():
    print(f"{row['rank']:<5} {row['sector_short']:<25} {row['labor_share']:<14.1f} "
          f"{row['va_growth_avg']:<12.2f} {row.get('capital_intensity', np.nan):<12.2f} "
          f"{row['vulnerability_index']:<8.1f}")


# ============================================================================
# 4. PUBLICATION-QUALITY FIGURES
# ============================================================================
print("\n" + "=" * 70)
print("GENERATING FIGURES")
print("=" * 70)

# Use short names for plotting
df_plot = df_summary.copy()
df_plot.index = [SECTOR_VERY_SHORT.get(s, s[:20]) for s in df_plot.index]

# --- Figure 4a: HEATMAP ---
print("\n[4a] Heatmap: Sectors x Indicators...")

heatmap_cols = ['labor_share', 'va_growth_avg', 'capital_intensity', 'vulnerability_index']
heatmap_labels = ['Participacion\nlaboral (%)', 'Crec. VA\npromedio (%)', 'Intensidad\nde capital', 'Indice\nvulnerabilidad']

df_heatmap = df_plot[heatmap_cols].copy()
df_heatmap.columns = heatmap_labels

# Normalize each column for coloring
df_heatmap_norm = df_heatmap.copy()
for c in df_heatmap_norm.columns:
    s = df_heatmap_norm[c].dropna()
    if len(s) > 0:
        mn, mx = s.min(), s.max()
        if mx > mn:
            df_heatmap_norm[c] = (df_heatmap_norm[c] - mn) / (mx - mn)

fig, ax = plt.subplots(figsize=(10, 8))

# Custom colormap
cmap = LinearSegmentedColormap.from_list('custom', ['#EBF5FB', MAINBLUE, '#1A5276'], N=256)

im = ax.imshow(df_heatmap_norm.values, cmap=cmap, aspect='auto', vmin=0, vmax=1)

ax.set_xticks(range(len(heatmap_labels)))
ax.set_xticklabels(heatmap_labels, fontsize=10)
ax.set_yticks(range(len(df_heatmap.index)))
ax.set_yticklabels(df_heatmap.index, fontsize=10)

# Annotate cells with actual values
for i in range(len(df_heatmap.index)):
    for j in range(len(heatmap_labels)):
        val = df_heatmap.iloc[i, j]
        norm_val = df_heatmap_norm.iloc[i, j]
        if pd.notna(val):
            color = 'white' if norm_val > 0.5 else 'black'
            text = f'{val:.1f}' if abs(val) < 100 else f'{val:.0f}'
            ax.text(j, i, text, ha='center', va='center', color=color, fontsize=9)

ax.set_title('Indicadores sectoriales de vulnerabilidad a la automatizacion\nColombia, 2005-2024',
             fontweight='bold', fontsize=13, pad=15)

plt.colorbar(im, ax=ax, shrink=0.6, label='Valor normalizado')
fig.tight_layout()
save_figure(fig, 'fig_heatmap_sectoral_indicators')


# --- Figure 4b: SCATTER PLOT ---
print("\n[4b] Scatter: Labor productivity vs labor cost share...")

fig, ax = plt.subplots(figsize=(10, 8))

# Use VA per sector as proxy for productivity level
# Bubble size = VA share
for i, (idx, row) in enumerate(df_plot.iterrows()):
    x = row['va_growth_avg']  # Productivity growth proxy
    y = row['labor_share']
    size = max(row['va_share_pct'] * 30, 60)  # Scale bubble size
    color = SECTOR_COLORS[i % len(SECTOR_COLORS)]

    ax.scatter(x, y, s=size, c=color, alpha=0.75, edgecolors='white', linewidth=1.5, zorder=5)

    # Label
    offset_x = 0.15
    offset_y = 1.0
    ax.annotate(idx, (x, y), fontsize=8, ha='left',
                xytext=(offset_x, offset_y), textcoords='offset points',
                fontweight='bold', color=color)

# Add quadrant lines at medians
x_med = df_plot['va_growth_avg'].median()
y_med = df_plot['labor_share'].median()
ax.axhline(y_med, color=LIGHTGRAY, linestyle='--', linewidth=0.8, alpha=0.7)
ax.axvline(x_med, color=LIGHTGRAY, linestyle='--', linewidth=0.8, alpha=0.7)

# Quadrant labels
ax.text(df_plot['va_growth_avg'].max() * 0.95, df_plot['labor_share'].max() * 0.98,
        'Alta participacion laboral\nAlto crecimiento',
        ha='right', va='top', fontsize=8, style='italic', color=DARKGRAY, alpha=0.7)
ax.text(df_plot['va_growth_avg'].min() * 1.05, df_plot['labor_share'].max() * 0.98,
        'Alta participacion laboral\nBajo crecimiento\n(MAYOR VULNERABILIDAD)',
        ha='left', va='top', fontsize=8, style='italic', color=ACCENTRED, fontweight='bold', alpha=0.8)

clean_style(ax,
            title='Crecimiento del valor agregado vs. participacion laboral por sector\nColombia, 2005-2024 (tamano = participacion en VA total)',
            xlabel='Crecimiento promedio anual del VA real (%)',
            ylabel='Participacion laboral estimada en VA (%)')

fig.tight_layout()
save_figure(fig, 'fig_scatter_productivity_vs_laborshare')


# --- Figure 4c: TIME SERIES PANEL - VA evolution by sector ---
print("\n[4c] Time series: Value added evolution by sector...")

fig, axes = plt.subplots(3, 4, figsize=(16, 12), sharex=True)
axes_flat = axes.flatten()

for i, s in enumerate(sector_names_12):
    if i >= 12:
        break
    ax = axes_flat[i]
    short_name = SECTOR_VERY_SHORT.get(s, s[:20])

    if s in df_va_const.columns:
        series = df_va_const[s].dropna()
        # Index to 2005 = 100
        base_val = series.iloc[0]
        indexed = series / base_val * 100
        ax.plot(indexed.index, indexed.values, color=SECTOR_COLORS[i % len(SECTOR_COLORS)],
                linewidth=2)
        ax.fill_between(indexed.index, 100, indexed.values,
                        alpha=0.15, color=SECTOR_COLORS[i % len(SECTOR_COLORS)])

    ax.axhline(100, color=LIGHTGRAY, linestyle='--', linewidth=0.7)
    ax.set_title(short_name, fontsize=10, fontweight='bold')
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    if i >= 8:
        ax.set_xlabel('Ano')
    if i % 4 == 0:
        ax.set_ylabel('Indice (2005=100)')

fig.suptitle('Evolucion del valor agregado real por sector\nColombia, 2005-2024 (Indice 2005 = 100)',
             fontweight='bold', fontsize=14, y=1.02)
fig.tight_layout()
save_figure(fig, 'fig_timeseries_va_by_sector')


# --- Figure 4d: BAR CHART - Automation Vulnerability Ranking ---
print("\n[4d] Bar chart: Automation vulnerability ranking...")

fig, ax = plt.subplots(figsize=(12, 7))

df_ranked = df_plot.sort_values('vulnerability_index', ascending=True)

colors_bar = []
for i, val in enumerate(df_ranked['vulnerability_index']):
    if val >= 60:
        colors_bar.append(ACCENTRED)
    elif val >= 45:
        colors_bar.append('#E67E22')
    elif val >= 30:
        colors_bar.append('#F39C12')
    else:
        colors_bar.append(MAINBLUE)

bars = ax.barh(range(len(df_ranked)), df_ranked['vulnerability_index'],
               color=colors_bar, edgecolor='white', linewidth=0.5, height=0.7)

ax.set_yticks(range(len(df_ranked)))
ax.set_yticklabels(df_ranked.index, fontsize=10)

# Add value labels on bars
for i, (val, bar) in enumerate(zip(df_ranked['vulnerability_index'], bars)):
    ax.text(val + 0.5, i, f'{val:.1f}', va='center', ha='left', fontsize=9, fontweight='bold')

clean_style(ax,
            title='Indice de vulnerabilidad a la automatizacion por sector\nColombia (mayor valor = mayor vulnerabilidad)',
            xlabel='Indice de vulnerabilidad (0-100)')

# Add legend
from matplotlib.patches import Patch
legend_elements = [
    Patch(facecolor=ACCENTRED, label='Alta vulnerabilidad (>=60)'),
    Patch(facecolor='#E67E22', label='Vulnerabilidad media-alta (45-60)'),
    Patch(facecolor='#F39C12', label='Vulnerabilidad media (30-45)'),
    Patch(facecolor=MAINBLUE, label='Vulnerabilidad baja (<30)'),
]
ax.legend(handles=legend_elements, loc='lower right', fontsize=9, framealpha=0.9)

fig.tight_layout()
save_figure(fig, 'fig_bar_vulnerability_ranking')


# --- Figure 4e: CAPITAL STOCK EVOLUTION ---
print("\n[4e] Capital stock evolution by sector...")

fig, ax = plt.subplots(figsize=(12, 7))

cap_sectors = [c for c in df_cap_net.columns if c != 'Total economía']
cap_data_plot = df_cap_net[cap_sectors].loc[2005:2023].dropna(how='all')

# Index to 2005 = 100
for i, s in enumerate(cap_sectors):
    series = cap_data_plot[s].dropna()
    base_val = series.iloc[0]
    indexed = series / base_val * 100
    ax.plot(indexed.index, indexed.values,
            color=SECTOR_COLORS[i % len(SECTOR_COLORS)],
            linewidth=2, label=s[:30], marker='o', markersize=3)

ax.axhline(100, color=LIGHTGRAY, linestyle='--', linewidth=0.7)

clean_style(ax,
            title='Evolucion del acervo de capital neto por sector\nColombia, 2005-2023 (Indice 2005 = 100)',
            xlabel='Ano',
            ylabel='Indice (2005 = 100)')

ax.legend(loc='upper left', fontsize=8, framealpha=0.9, ncol=2)
fig.tight_layout()
save_figure(fig, 'fig_capital_stock_evolution')


# --- ADDITIONAL FIGURE: Aggregate labor share over time ---
print("\n[Extra] Aggregate labor share over time...")

fig, ax = plt.subplots(figsize=(10, 5))

ax.plot(df_labor_share_agg.index, df_labor_share_agg.values,
        color=MAINBLUE, linewidth=2.5, marker='o', markersize=4)
ax.fill_between(df_labor_share_agg.index, df_labor_share_agg.values,
                alpha=0.15, color=MAINBLUE)

ax.axhline(df_labor_share_agg.mean(), color=ACCENTRED, linestyle='--',
           linewidth=1, alpha=0.7, label=f'Promedio: {df_labor_share_agg.mean():.1f}%')

clean_style(ax,
            title='Participacion de la remuneracion de asalariados en el valor agregado bruto\nColombia, 2014-2024',
            xlabel='Ano',
            ylabel='Participacion laboral (% del VA)')

ax.legend(loc='lower right', fontsize=10)
ax.yaxis.set_major_formatter(mticker.FormatStrFormatter('%.1f%%'))
fig.tight_layout()
save_figure(fig, 'fig_aggregate_labor_share')


# --- ADDITIONAL FIGURE: TFP by sector ---
print("\n[Extra] TFP growth by sector...")

fig, ax = plt.subplots(figsize=(12, 7))

# Average TFP by sector over the period
tfp_avg = {}
for s in PTF_SECTORS:
    if s in df_tfp.columns:
        vals = df_tfp[s].dropna()
        if len(vals) > 0:
            tfp_avg[s] = vals.mean()

tfp_sorted = sorted(tfp_avg.items(), key=lambda x: x[1])
sectors_tfp = [x[0][:30] for x in tfp_sorted]
values_tfp = [x[1] for x in tfp_sorted]

colors_tfp = [ACCENTRED if v < 0 else MAINBLUE for v in values_tfp]

ax.barh(range(len(sectors_tfp)), values_tfp, color=colors_tfp,
        edgecolor='white', linewidth=0.5, height=0.6)
ax.set_yticks(range(len(sectors_tfp)))
ax.set_yticklabels(sectors_tfp, fontsize=10)

for i, v in enumerate(values_tfp):
    ax.text(v + (0.05 if v >= 0 else -0.05), i,
            f'{v:.2f}%', va='center', ha='left' if v >= 0 else 'right',
            fontsize=9, fontweight='bold')

ax.axvline(0, color='black', linewidth=0.8)

clean_style(ax,
            title='Productividad Total de los Factores promedio por sector\nColombia, 2005-2024 (crecimiento anual %)',
            xlabel='PTF promedio anual (%)')

fig.tight_layout()
save_figure(fig, 'fig_tfp_by_sector')


# ============================================================================
# 5. SAVE RESULTS
# ============================================================================
print("\n" + "=" * 70)
print("SAVING RESULTS")
print("=" * 70)

# Prepare final output table
df_output = df_summary.copy()
df_output = df_output.reset_index()
df_output = df_output.rename(columns={
    'sector': 'sector_full',
    'sector_short': 'sector',
    'va_share_pct': 'va_share_pct_2024',
    'va_current_2024': 'va_corrientes_2024_mmCOP',
    'va_const_2024': 'va_constantes_2024_mmCOP',
    'va_growth_avg': 'crecimiento_va_promedio_pct',
    'va_growth_recent': 'crecimiento_va_reciente_pct',
    'labor_share': 'participacion_laboral_pct',
    'capital_intensity': 'intensidad_capital_ratio',
    'vulnerability_index': 'indice_vulnerabilidad',
    'rank': 'ranking_vulnerabilidad',
})

cols_out = [
    'ranking_vulnerabilidad', 'sector', 'sector_full',
    'va_share_pct_2024', 'va_corrientes_2024_mmCOP', 'va_constantes_2024_mmCOP',
    'crecimiento_va_promedio_pct', 'crecimiento_va_reciente_pct',
    'participacion_laboral_pct', 'intensidad_capital_ratio',
    'labor_share_norm', 'prod_growth_norm', 'cap_intensity_norm',
    'indice_vulnerabilidad',
]
df_output = df_output[[c for c in cols_out if c in df_output.columns]]

output_path = os.path.join(OUT_DIR, 'sectoral_analysis_results.csv')
df_output.to_csv(output_path, index=False, encoding='utf-8-sig')
print(f"  Results saved to: {output_path}")

# Also save the VA time series
va_ts_path = os.path.join(OUT_DIR, 'va_timeseries_constant_prices.csv')
df_va_const_out = df_va_const.copy()
df_va_const_out.columns = [SECTOR_VERY_SHORT.get(c, c[:30]) for c in df_va_const_out.columns]
df_va_const_out.to_csv(va_ts_path, encoding='utf-8-sig')
print(f"  VA time series saved to: {va_ts_path}")

# Save TFP time series
tfp_ts_path = os.path.join(OUT_DIR, 'tfp_timeseries_by_sector.csv')
df_tfp.to_csv(tfp_ts_path, encoding='utf-8-sig')
print(f"  TFP time series saved to: {tfp_ts_path}")

# Save capital stock time series
cap_ts_path = os.path.join(OUT_DIR, 'capital_stock_timeseries.csv')
df_cap_net.to_csv(cap_ts_path, encoding='utf-8-sig')
print(f"  Capital stock saved to: {cap_ts_path}")


# ============================================================================
# 6. SUMMARY OUTPUT
# ============================================================================
print("\n" + "=" * 70)
print("SUMMARY OF KEY FINDINGS")
print("=" * 70)

print(f"\nTotal Value Added (current prices, {latest_year}): {total_va:,.0f} miles de millones COP")
if 'PIB' in df_va_current.columns:
    print(f"Total GDP ({latest_year}): {df_va_current['PIB'].iloc[-1]:,.0f} miles de millones COP")
else:
    print("Total GDP: not extracted (PIB row not found in expected range)")

if len(df_labor_share_agg) > 0:
    print(f"\nAggregate labor share: {df_labor_share_agg.iloc[-1]:.1f}%")
else:
    print("\nAggregate labor share: not available")

print("\nTop 5 most vulnerable sectors to automation:")
top5 = df_summary.head(5)
for _, row in top5.iterrows():
    print(f"  {row['rank']}. {row['sector_short']} (index={row['vulnerability_index']:.1f}, "
          f"labor share={row['labor_share']:.1f}%, VA growth={row['va_growth_avg']:.2f}%)")

print("\nBottom 3 (least vulnerable):")
bottom3 = df_summary.tail(3)
for _, row in bottom3.iterrows():
    print(f"  {row['rank']}. {row['sector_short']} (index={row['vulnerability_index']:.1f}, "
          f"labor share={row['labor_share']:.1f}%, VA growth={row['va_growth_avg']:.2f}%)")

print("\nNOTES:")
print("  - Labor share by sector is estimated using structural priors and")
print("    institutional sector composition from DANE's Cuadro cruzado.")
print("    DANE does not publish sector-level labor compensation directly in")
print("    the available national accounts data; this is a known limitation.")
print("  - Capital intensity for sectors without direct capital stock data")
print("    is estimated by proportional allocation from 9-sector KLEMS groupings.")
print("  - Vulnerability index weights: labor share (40%), productivity growth (35%),")
print("    capital intensity inverse (25%).")
print("  - TFP data uses CIIU 3.0/KLEMS classification (9 sectors).")
print("  - VA data uses CIIU 4.0 classification (12 sector groupings).")

print("\n" + "=" * 70)
print("SECTORAL ANALYSIS COMPLETE")
print("=" * 70)
